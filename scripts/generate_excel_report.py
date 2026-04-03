"""
Generates the RetainIQ VoC Research Brief — Excel Report
4 tabs: VoC Summary | Pain Point Analysis | Win-Back Opportunity | Recommendations
Author: Alisha Sharma
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_csv("/home/claude/voice-of-customer-voc-analysis/data/voc_analysis_output.csv")

NAVY   = "1F3864"; BLUE  = "2E75B6"; LBLUE = "DEEAF1"
GREEN  = "E2EFDA"; RED   = "FCE4D6"; YELLOW= "FFF2CC"
LGRAY  = "F2F2F2"; WHITE = "FFFFFF"; DGRAY = "333333"
ORANGE = "F4B942"

PAIN_COLORS = {
    "Product Quality":       "FCE4D6",
    "Customer Support":      "FCE4D6",
    "Pricing & Value":       "FFEB9C",
    "Onboarding & Usability":"FFF2CC",
    "Competitor Switch":     "DEEAF1",
}

def fill(c):   return PatternFill("solid", fgColor=c)
def fnt(c="FFFFFF", bold=True, sz=9): return Font(color=c, bold=bold, size=sz, name="Calibri")
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def lft(wrap=False): return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)

def sc(ws, r, c, v, f=None, fn=None, al=None):
    cell = ws.cell(row=r, column=c, value=v)
    if f:  cell.fill      = fill(f)
    if fn: cell.font      = fn
    if al: cell.alignment = al
    cell.border = bdr()
    return cell

wb = openpyxl.Workbook()

churned  = df[df["churn_status"] == "Churned"]
at_risk  = df[df["churn_status"] == "At-Risk"]
retained = df[df["churn_status"] == "Retained"]

mrr_lost    = churned["monthly_recurring_revenue"].sum()
mrr_risk    = at_risk["monthly_recurring_revenue"].sum()
winback_df  = churned[churned["win_back_receptive"] == "Yes"]
winback_mrr = winback_df["monthly_recurring_revenue"].sum()

PAIN_ORDER = churned["primary_pain_point"].value_counts().index.tolist()

# ═══════════════════════════════════════════════════════════════
# SHEET 1 — VOC SUMMARY
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "VoC Summary"
ws1.sheet_view.showGridLines = False
for i, w in enumerate([28,18,18,18,18,18], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.row_dimensions[1].height = 35
ws1.merge_cells("A1:F1")
c = ws1.cell(row=1,column=1,value="RetainIQ — Voice of Customer Research Brief")
c.fill=fill(NAVY); c.font=Font(color="FFFFFF",bold=True,size=15,name="Calibri"); c.alignment=ctr()

ws1.row_dimensions[2].height = 18
ws1.merge_cells("A2:F2")
c = ws1.cell(row=2,column=1,value="Customer Churn & Retention Study  |  600 Respondents  |  Research & Insights Team")
c.fill=fill(BLUE); c.font=fnt(sz=9); c.alignment=ctr()

# Key metrics
ws1.row_dimensions[4].height = 14
ws1.merge_cells("A4:F4")
c = ws1.cell(row=4,column=1,value="STUDY OVERVIEW")
c.fill=fill(BLUE); c.font=fnt(sz=10); c.alignment=lft()

kpis = [
    ("Total Respondents",      "600"),
    ("Churned Customers",      f"{len(churned)} ({round(len(churned)/len(df)*100,1)}%)"),
    ("At-Risk Customers",      f"{len(at_risk)} ({round(len(at_risk)/len(df)*100,1)}%)"),
    ("MRR Lost (Churned)",     f"${mrr_lost:,}/month"),
    ("MRR At Risk",            f"${mrr_risk:,}/month"),
    ("Win-Back Opportunity",   f"${winback_mrr:,}/month (${winback_mrr*12:,} ARR)"),
    ("Win-Back Receptive Rate",f"{round(len(winback_df)/len(churned)*100,1)}% of churned"),
    ("Avg Exit Satisfaction",  f"{round(churned['satisfaction_at_exit'].mean(),2)} / 5.0"),
    ("Top Pain Point",         PAIN_ORDER[0] if PAIN_ORDER else "N/A"),
    ("Avg Repurchase Intent",  f"{round(churned['repurchase_intent'].mean(),2)} / 5.0"),
]

for idx, (label, val) in enumerate(kpis):
    col = (idx % 2) * 3 + 1
    row = 5 + (idx // 2)
    ws1.row_dimensions[row].height = 22
    c1 = ws1.cell(row=row,column=col,value=label)
    c1.fill=fill(LGRAY); c1.font=fnt(DGRAY,True,9); c1.alignment=lft(); c1.border=bdr()
    ws1.merge_cells(f"{get_column_letter(col+1)}{row}:{get_column_letter(col+2)}{row}")
    c2 = ws1.cell(row=row,column=col+1,value=val)
    bg = RED if "Lost" in label or "Churned" in label else (YELLOW if "Risk" in label else (GREEN if "Win-Back" in label else WHITE))
    c2.fill=fill(bg); c2.font=fnt("000000",False,9); c2.alignment=ctr(); c2.border=bdr()

# Churn by tier
ws1.row_dimensions[11].height = 14
ws1.merge_cells("A11:F11")
c = ws1.cell(row=11,column=1,value="CHURN STATUS BY SUBSCRIPTION TIER")
c.fill=fill(BLUE); c.font=fnt(sz=10); c.alignment=lft()

tier_hdrs = ["Subscription Tier","Total","Churned","At-Risk","Retained","Churn Rate"]
for i,h in enumerate(tier_hdrs,1):
    ws1.row_dimensions[12].height = 20
    c = ws1.cell(row=12,column=i,value=h)
    c.fill=fill(LBLUE); c.font=fnt("000000",True,9); c.alignment=ctr(); c.border=bdr()

for r, tier in enumerate(["Basic","Pro","Enterprise"],13):
    t  = df[df["subscription_tier"]==tier]
    ch = t[t["churn_status"]=="Churned"]
    ar = t[t["churn_status"]=="At-Risk"]
    re = t[t["churn_status"]=="Retained"]
    cr = round(len(ch)/len(t)*100,1)
    ws1.row_dimensions[r].height = 20
    bg = LGRAY if r%2==0 else WHITE
    for i,v in enumerate([tier,len(t),len(ch),len(ar),len(re),f"{cr}%"],1):
        c = ws1.cell(row=r,column=i,value=v)
        fc = RED if i==3 else (YELLOW if i==4 else (GREEN if i==5 else bg))
        c.fill=fill(fc); c.font=fnt("000000",i==1,9)
        c.alignment=lft() if i==1 else ctr(); c.border=bdr()

# ═══════════════════════════════════════════════════════════════
# SHEET 2 — PAIN POINT ANALYSIS
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Pain Point Analysis")
ws2.sheet_view.showGridLines = False
for i,w in enumerate([26,12,12,14,16,16,16,16],1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:H1")
c = ws2.cell(row=1,column=1,value="Pain Point Analysis — Churned Customers (n=271)")
c.fill=fill(NAVY); c.font=Font(color="FFFFFF",bold=True,size=13,name="Calibri"); c.alignment=ctr()
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:H2")
c = ws2.cell(row=2,column=1,value="Sorted by severity — lowest exit satisfaction first")
c.fill=fill(BLUE); c.font=fnt(sz=9); c.alignment=ctr()

ws2.merge_cells("A4:H4")
c = ws2.cell(row=4,column=1,value="PAIN POINT FREQUENCY & SEVERITY")
c.fill=fill(BLUE); c.font=fnt(sz=10); c.alignment=lft()

pp_hdrs = ["Pain Point","Customers","% of Churned","Avg Satisfaction","Avg Repurchase","Avg Sentiment","MRR Lost","Neg Feedback%"]
for i,h in enumerate(pp_hdrs,1):
    ws2.row_dimensions[5].height = 22
    c = ws2.cell(row=5,column=i,value=h)
    c.fill=fill(LBLUE); c.font=fnt("000000",True,9); c.alignment=ctr(True); c.border=bdr()

pain_data = churned.groupby("primary_pain_point").agg(
    count=("customer_id","count"),
    avg_sat=("satisfaction_at_exit","mean"),
    avg_rep=("repurchase_intent","mean"),
    avg_sent=("sentiment_score","mean"),
    mrr=("monthly_recurring_revenue","sum"),
    neg_pct=("sentiment_category", lambda x: round((x=="Negative").sum()/len(x)*100,1))
).round(2).sort_values("avg_sat")

for r, (pain, row) in enumerate(pain_data.iterrows(), 6):
    ws2.row_dimensions[r].height = 22
    bg = PAIN_COLORS.get(pain, WHITE)
    pct = round(row["count"]/len(churned)*100,1)
    vals = [pain, int(row["count"]), f"{pct}%",
            f"{row['avg_sat']} / 5",
            f"{row['avg_rep']} / 5",
            round(row["avg_sent"],3),
            f"${int(row['mrr']):,}",
            f"{row['neg_pct']}%"]
    for i,v in enumerate(vals,1):
        c = ws2.cell(row=r,column=i,value=v)
        c.fill=fill(bg if i==1 else (LGRAY if r%2==0 else WHITE))
        c.font=fnt("000000",i==1,9)
        c.alignment=lft() if i==1 else ctr(); c.border=bdr()

# Sentiment by churn status
ws2.row_dimensions[13].height = 14
ws2.merge_cells("A13:H13")
c = ws2.cell(row=13,column=1,value="EXIT FEEDBACK SENTIMENT BY CHURN STATUS (VADER Analysis)")
c.fill=fill(BLUE); c.font=fnt(sz=10); c.alignment=lft()

sent_hdrs = ["Churn Status","Total","Positive","Neutral","Negative","% Negative","Avg Score"]
for i,h in enumerate(sent_hdrs,1):
    ws2.row_dimensions[14].height = 20
    c = ws2.cell(row=14,column=i,value=h)
    c.fill=fill(LBLUE); c.font=fnt("000000",True,9); c.alignment=ctr(); c.border=bdr()

status_colors = {"Churned":RED,"At-Risk":YELLOW,"Retained":GREEN}
for r,status in enumerate(["Churned","At-Risk","Retained"],15):
    s = df[df["churn_status"]==status]
    pos=len(s[s["sentiment_category"]=="Positive"])
    neu=len(s[s["sentiment_category"]=="Neutral"])
    neg=len(s[s["sentiment_category"]=="Negative"])
    ws2.row_dimensions[r].height = 20
    bg = status_colors[status]
    vals = [status,len(s),pos,neu,neg,f"{round(neg/len(s)*100,1)}%",round(s["sentiment_score"].mean(),3)]
    for i,v in enumerate(vals,1):
        c = ws2.cell(row=r,column=i,value=v)
        c.fill=fill(bg if i==1 else (LGRAY if r%2==0 else WHITE))
        c.font=fnt("000000",i==1,9); c.alignment=ctr() if i>1 else lft(); c.border=bdr()

# ═══════════════════════════════════════════════════════════════
# SHEET 3 — WIN-BACK OPPORTUNITY
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Win-Back Opportunity")
ws3.sheet_view.showGridLines = False
for i,w in enumerate([26,16,16,18,18,16],1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:F1")
c = ws3.cell(row=1,column=1,value="Win-Back & Revenue Recovery Opportunity")
c.fill=fill(NAVY); c.font=Font(color="FFFFFF",bold=True,size=13,name="Calibri"); c.alignment=ctr()
ws3.row_dimensions[1].height = 30

# Summary box
ws3.merge_cells("A3:F3")
c = ws3.cell(row=3,column=1,value="WIN-BACK OPPORTUNITY SUMMARY")
c.fill=fill(BLUE); c.font=fnt(sz=10); c.alignment=lft()

wb_kpis = [
    ("Total Churned Customers",    len(churned)),
    ("Win-Back Receptive",         f"{len(winback_df)} ({round(len(winback_df)/len(churned)*100,1)}%)"),
    ("Recoverable MRR",            f"${winback_mrr:,}/month"),
    ("Recoverable ARR",            f"${winback_mrr*12:,}/year"),
    ("Highest Value Tier",         "Enterprise"),
    ("At-Risk MRR (Preventable)", f"${mrr_risk:,}/month"),
]
for idx,(label,val) in enumerate(wb_kpis):
    col = (idx%2)*3+1
    row = 4+(idx//2)
    ws3.row_dimensions[row].height = 22
    c1 = ws3.cell(row=row,column=col,value=label)
    c1.fill=fill(LGRAY); c1.font=fnt(DGRAY,True,9); c1.alignment=lft(); c1.border=bdr()
    ws3.merge_cells(f"{get_column_letter(col+1)}{row}:{get_column_letter(col+2)}{row}")
    c2 = ws3.cell(row=row,column=col+1,value=val)
    c2.fill=fill(GREEN); c2.font=fnt("000000",True,9); c2.alignment=ctr(); c2.border=bdr()

# Win-back by tier
ws3.row_dimensions[8].height = 14
ws3.merge_cells("A8:F8")
c = ws3.cell(row=8,column=1,value="WIN-BACK OPPORTUNITY BY SUBSCRIPTION TIER")
c.fill=fill(BLUE); c.font=fnt(sz=10); c.alignment=lft()

wb_hdrs = ["Tier","Churned","Win-Back Count","Win-Back Rate","Recoverable MRR","Recoverable ARR"]
for i,h in enumerate(wb_hdrs,1):
    ws3.row_dimensions[9].height = 20
    c = ws3.cell(row=9,column=i,value=h)
    c.fill=fill(LBLUE); c.font=fnt("000000",True,9); c.alignment=ctr(True); c.border=bdr()

for r,tier in enumerate(["Enterprise","Pro","Basic"],10):
    t   = churned[churned["subscription_tier"]==tier]
    wb_ = t[t["win_back_receptive"]=="Yes"]
    mrr = wb_["monthly_recurring_revenue"].sum()
    ws3.row_dimensions[r].height = 22
    vals=[tier,len(t),len(wb_),f"{round(len(wb_)/len(t)*100,1)}%",f"${mrr:,}",f"${mrr*12:,}"]
    for i,v in enumerate(vals,1):
        bg = GREEN if i>=5 else (LGRAY if r%2==0 else WHITE)
        c = ws3.cell(row=r,column=i,value=v)
        c.fill=fill(bg); c.font=fnt("000000",i==1,9)
        c.alignment=lft() if i==1 else ctr(); c.border=bdr()

# Win-back by pain point
ws3.row_dimensions[15].height = 14
ws3.merge_cells("A15:F15")
c = ws3.cell(row=15,column=1,value="WIN-BACK RECEPTIVENESS BY PAIN POINT")
c.fill=fill(BLUE); c.font=fnt(sz=10); c.alignment=lft()

wbp_hdrs = ["Pain Point","Churned","Win-Back Receptive","Win-Back Rate","Recoverable MRR","Avg Repurchase Intent"]
for i,h in enumerate(wbp_hdrs,1):
    ws3.row_dimensions[16].height = 20
    c = ws3.cell(row=16,column=i,value=h)
    c.fill=fill(LBLUE); c.font=fnt("000000",True,9); c.alignment=ctr(True); c.border=bdr()

for r,(pain,row_data) in enumerate(
    churned.groupby("primary_pain_point").apply(
        lambda x: pd.Series({
            "count": len(x),
            "wb_count": (x["win_back_receptive"]=="Yes").sum(),
            "wb_rate": round((x["win_back_receptive"]=="Yes").sum()/len(x)*100,1),
            "wb_mrr": x[x["win_back_receptive"]=="Yes"]["monthly_recurring_revenue"].sum(),
            "avg_rep": round(x["repurchase_intent"].mean(),2)
        })
    ).sort_values("wb_rate",ascending=False).iterrows(),
    17):
    ws3.row_dimensions[r].height = 22
    bg = PAIN_COLORS.get(pain,WHITE)
    vals=[pain,int(row_data["count"]),int(row_data["wb_count"]),
          f"{row_data['wb_rate']}%",f"${int(row_data['wb_mrr']):,}",
          f"{row_data['avg_rep']} / 5"]
    for i,v in enumerate(vals,1):
        c = ws3.cell(row=r,column=i,value=v)
        c.fill=fill(bg if i==1 else (LGRAY if r%2==0 else WHITE))
        c.font=fnt("000000",i==1,9); c.alignment=lft() if i==1 else ctr(); c.border=bdr()

# ═══════════════════════════════════════════════════════════════
# SHEET 4 — RECOMMENDATIONS
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Retention Recommendations")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 5
ws4.column_dimensions["B"].width = 78

ws4.merge_cells("A1:B1")
c = ws4.cell(row=1,column=1,value="RetainIQ — VoC Retention Recommendations")
c.fill=fill(NAVY); c.font=Font(color="FFFFFF",bold=True,size=13,name="Calibri"); c.alignment=ctr()
ws4.row_dimensions[1].height = 30

recs = [
    (None,"IMMEDIATE ACTIONS — HIGH IMPACT",BLUE),
    ("1","Product Quality is the #1 churn driver (32.5% of churned customers, $25,039 MRR lost). Prioritize a bug resolution sprint and establish a formal product feedback loop with a 72-hour response SLA.","FCE4D6"),
    ("2","Customer Support drives 20.3% of churn. Implement tiered SLA targets: 4-hour response for Pro, 1-hour for Enterprise. Assign dedicated CSMs to all Enterprise accounts immediately.","FCE4D6"),
    ("3","Launch a proactive win-back campaign targeting the 36.9% of churned customers who are receptive. Enterprise win-back alone represents $147,408 in recoverable ARR.","C6EFCE"),
    (None,"PRICING & VALUE IMPROVEMENTS",BLUE),
    ("4","Pricing & Value ranks #2 in churn drivers (23.2% of churned customers). Introduce a flexible month-to-month downgrade path for at-risk Basic customers to reduce full churn.","FFEB9C"),
    ("5","Communicate price increases minimum 60 days before renewal with a documented feature improvement summary. Lack of transparency is a leading driver of negative exit sentiment.","FFEB9C"),
    ("6","Consider a loyalty pricing tier for customers with 24+ months tenure — this cohort represents significant MRR and has higher win-back potential if proactively managed.","FFEB9C"),
    (None,"ONBOARDING & RETENTION IMPROVEMENTS",BLUE),
    ("7","Onboarding failure accounts for 13.7% of churn — highest among early-tenure customers (0-6 months). Launch a 30-60-90 day structured onboarding program with milestone check-ins.","FFF2CC"),
    ("8","Assign a dedicated Customer Success Manager to all new Pro and Enterprise accounts for the first 90 days. Early engagement is the single highest-ROI retention investment available.","FFF2CC"),
    ("9","Build an in-product health score dashboard flagging at-risk accounts based on login frequency, support ticket volume, and feature adoption rate.","FFF2CC"),
    (None,"AT-RISK CUSTOMER PRIORITIZATION",BLUE),
    ("10","Prioritize outreach to 188 At-Risk customers before renewal — preventing churn is 5x more cost-effective than winning back churned customers. Begin with highest MRR accounts.","BDD7EE"),
    ("11","Technology and Healthcare verticals show the highest churn rates (50.9% and 50.0%). Build industry-specific customer success playbooks addressing the unique workflows of these segments.","BDD7EE"),
]

row = 3
for num, text, bg in recs:
    ws4.row_dimensions[row].height = 40
    if num is None:
        ws4.merge_cells(f"A{row}:B{row}")
        c = ws4.cell(row=row,column=1,value=text)
        c.fill=fill(bg); c.font=fnt(sz=10); c.alignment=lft(); c.border=bdr()
    else:
        c1 = ws4.cell(row=row,column=1,value=num)
        c1.fill=fill(LGRAY); c1.font=fnt(DGRAY,True,9); c1.alignment=ctr(); c1.border=bdr()
        c2 = ws4.cell(row=row,column=2,value=text)
        c2.fill=fill(bg); c2.font=fnt("000000",False,9); c2.alignment=lft(True); c2.border=bdr()
    row += 1

wb.save("/home/claude/voice-of-customer-voc-analysis/output/RetainIQ_VoC_Research_Brief.xlsx")
print("Excel report saved.")
